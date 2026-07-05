#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
知识库 Excel 导入脚本（Python版）

使用方式：
    python3 scripts/import_knowledge.py [excel文件路径]

示例：
    python3 scripts/import_knowledge.py data/knowledge.xlsx
    python3 scripts/import_knowledge.py /Users/zouxuefei/Desktop/知识库.xlsx

依赖安装：
    pip3 install openpyxl elasticsearch
"""

import sys
import os
from datetime import datetime

# 检查依赖
try:
    import openpyxl
except ImportError:
    print("❌ 缺少依赖: openpyxl")
    print("   安装命令: pip3 install openpyxl")
    sys.exit(1)

try:
    from elasticsearch import Elasticsearch
except ImportError:
    print("❌ 缺少依赖: elasticsearch")
    print("   安装命令: pip3 install elasticsearch")
    sys.exit(1)


# ============ 配置 ============

ES_HOST = "http://localhost:9200"
INDEX_NAME = "crm_knowledge"

# 有效分类
VALID_CATEGORIES = ["PLATFORM", "TICKET", "FAQ", "SERVICE", "BENEFITS", "SCRIPT", "SYSTEM"]


def connect_es():
    """连接 ES"""
    try:
        es = Elasticsearch(ES_HOST)
        if es.ping():
            print(f"✅ 已连接 ES: {ES_HOST}")
            return es
        else:
            print(f"❌ 无法连接 ES: {ES_HOST}")
            sys.exit(1)
    except Exception as e:
        print(f"❌ ES 连接失败: {e}")
        sys.exit(1)


def read_excel(file_path):
    """读取 Excel 文件"""
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        sys.exit(1)

    print(f"📖 读取 Excel: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    print(f"   表头: {headers}")

    # 解析数据行
    articles = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过空行
        if not row or not row[0]:
            continue

        # 构建数据字典
        data = dict(zip(headers, row))

        article = parse_row(data, row_idx)
        if article:
            articles.append(article)

    print(f"   读取完成: {len(articles)} 条有效数据")
    return articles


def parse_row(data, row_idx):
    """解析单行数据"""
    title = data.get("标题")
    content = data.get("内容")

    # 校验必填字段
    if not title or not content:
        print(f"   ⚠️  第{row_idx}行: 标题或内容为空，跳过")
        return None

    # 处理分类
    category = data.get("分类", "FAQ")
    if category:
        category = str(category).strip().upper()
    if category not in VALID_CATEGORIES:
        print(f"   ⚠️  第{row_idx}行: 无效分类 '{category}'，使用默认值 FAQ")
        category = "FAQ"

    # 处理标签
    tags = []
    tags_str = data.get("标签")
    if tags_str:
        # 支持中英文逗号分隔
        tags = [t.strip() for t in str(tags_str).replace("，", ",").split(",") if t.strip()]

    # 处理关键词
    keywords = data.get("关键词")
    if keywords:
        keywords = str(keywords).strip()

    # 处理优先级
    priority = data.get("优先级")
    if priority is None:
        priority = 5
    try:
        priority = int(priority)
    except (ValueError, TypeError):
        priority = 5

    return {
        "title": str(title).strip(),
        "content": str(content).strip(),
        "category": category,
        "tags": tags,
        "keywords": keywords,
        "priority": priority,
        "viewCount": 0,
        "status": 1,
        "createTime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updateTime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "createBy": "python-import"
    }


def create_index(es):
    """创建 ES 索引（如果不存在）"""
    if es.indices.exists(index=INDEX_NAME):
        print(f"ℹ️  索引已存在: {INDEX_NAME}")
        return

    print(f"📝 创建索引: {INDEX_NAME}")

    mapping = {
        "mappings": {
            "properties": {
                "title": {"type": "text", "analyzer": "ik_max_word", "search_analyzer": "ik_smart"},
                "content": {"type": "text", "analyzer": "ik_smart"},
                "category": {"type": "keyword"},
                "tags": {"type": "keyword"},
                "keywords": {"type": "text", "analyzer": "ik_max_word"},
                "priority": {"type": "integer"},
                "viewCount": {"type": "integer"},
                "status": {"type": "integer"},
                "createTime": {"type": "date", "format": "yyyy-MM-dd HH:mm:ss||yyyy-MM-dd||epoch_millis||strict_date_optional_time"},
                "updateTime": {"type": "date", "format": "yyyy-MM-dd HH:mm:ss||yyyy-MM-dd||epoch_millis||strict_date_optional_time"},
                "createBy": {"type": "keyword"}
            }
        }
    }

    es.indices.create(index=INDEX_NAME, body=mapping)
    print(f"✅ 索引创建成功")


def import_to_es(es, articles, overwrite=False):
    """批量导入到 ES"""
    if not articles:
        print("ℹ️  没有数据需要导入")
        return 0

    print(f"📤 开始导入 {len(articles)} 条数据到 ES...")

    # 如果覆盖模式，先删除同名数据
    if overwrite:
        for article in articles:
            # 按标题查询并删除
            query = {
                "query": {
                    "match": {
                        "title": article["title"]
                    }
                }
            }
            try:
                es.delete_by_query(index=INDEX_NAME, body=query)
            except Exception:
                pass

    # 批量写入
    success_count = 0
    fail_count = 0

    for i, article in enumerate(articles, 1):
        try:
            es.index(index=INDEX_NAME, body=article)
            success_count += 1
            print(f"   ✅ [{i}/{len(articles)}] {article['title']}")
        except Exception as e:
            fail_count += 1
            print(f"   ❌ [{i}/{len(articles)}] {article['title']}: {e}")

    print(f"\n📊 导入完成: 成功 {success_count} 条, 失败 {fail_count} 条")
    return success_count


def show_stats(es):
    """显示统计信息"""
    print("\n=== 导入后统计 ===")

    # 文档总数
    count = es.count(index=INDEX_NAME)["count"]
    print(f"📊 知识总数: {count}")

    # 分类分布
    agg_query = {
        "size": 0,
        "aggs": {
            "categories": {
                "terms": {"field": "category", "size": 10}
            }
        }
    }
    result = es.search(index=INDEX_NAME, body=agg_query)
    buckets = result["aggregations"]["categories"]["buckets"]

    print("📁 分类分布:")
    for bucket in buckets:
        print(f"   {bucket['key']}: {bucket['doc_count']}条")


def main():
    """主函数"""
    # 获取文件路径
    if len(sys.argv) < 2:
        # 默认路径
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_dir = os.path.dirname(script_dir)
        file_path = os.path.join(project_dir, "data", "knowledge.xlsx")
    else:
        file_path = sys.argv[1]

    print("=" * 50)
    print("🚀 知识库 Excel 导入工具 (Python版)")
    print("=" * 50)

    # 1. 连接 ES
    es = connect_es()

    # 2. 创建索引
    create_index(es)

    # 3. 读取 Excel
    articles = read_excel(file_path)

    # 4. 导入 ES
    success_count = import_to_es(es, articles, overwrite=True)

    # 5. 显示统计
    if success_count > 0:
        show_stats(es)

    print("\n✅ 全部完成！")


if __name__ == "__main__":
    main()
